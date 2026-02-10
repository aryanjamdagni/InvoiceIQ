import os
import re
import ast
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from config.settings import EXCEL_FILE
from core.schema_manager import load_schemas

def sanitize_sheet_name(name):
    if not name: return "Unknown_Vendor"
    clean_name = re.sub(r'[\[\]:*?/\\]', '', str(name))
    return clean_name[:25]

def format_excel_sheet(file_path, sheet_name):
    try:
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            FILL_RED    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            FILL_BLUE   = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            FILL_WHITE  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            HEADER_COLOR_MAP = {
                "Asset Category": FILL_RED,
                "Asset Class": FILL_RED,
                "Total Base Price": FILL_RED,
                "Total Tax": FILL_RED,
                "Total Purchase Price": FILL_RED,
                
                "Vendor Name": FILL_BLUE,
                "Date of Entry": FILL_BLUE,
                "Data Entry Done By": FILL_BLUE,
                "Customer Name": FILL_BLUE,
                
                "Client Code": FILL_YELLOW,
                "PO Number": FILL_YELLOW,
                "Invoice Status ( Proforma/ Final)": FILL_YELLOW,
                "Invoice No": FILL_YELLOW,
                "Invoice Date": FILL_YELLOW,
                "Invoice Received Date": FILL_YELLOW,
                "Vendor Code": FILL_YELLOW,
                "Vendor State": FILL_YELLOW,
                "Asset Type": FILL_YELLOW,
                "RS #": FILL_YELLOW,
                "Asset Description": FILL_YELLOW,
                "Asset Serial Number": FILL_YELLOW,
                "Qty": FILL_YELLOW,
                "UOM": FILL_YELLOW,
                "Deliver Address": FILL_YELLOW,
                "Delivery Location": FILL_YELLOW,
                "Delivery State": FILL_YELLOW,
                "Currency Code": FILL_YELLOW,
                "HSN/SAC": FILL_YELLOW,
                "Code": FILL_YELLOW,
                "Base Inv Amt (Excl Tax) - Material": FILL_YELLOW,
                "Base Inv Amt (Excl Tax) - Labour": FILL_YELLOW,
                "Reverse Charge": FILL_YELLOW,
                "Others": FILL_YELLOW,
                "CGST": FILL_YELLOW,
                "SGST": FILL_YELLOW,
                "IGST": FILL_YELLOW,
                "BCD": FILL_YELLOW,
                "% of Total Purchase Price being scheduled": FILL_YELLOW,
                "Waybill Inward Required ": FILL_YELLOW,
                "Waybill Outward Required ": FILL_YELLOW,
                "TDS Section": FILL_YELLOW,
                "TDS Base Value": FILL_YELLOW,
                "Vend Inv Type": FILL_YELLOW,
                "Remarks": FILL_YELLOW,
                "RAPL Billing State ": FILL_YELLOW,
                
                "Asset Make": FILL_WHITE,
                "Asset Model": FILL_WHITE,
                "LBT Applicable (Yes/No)": FILL_WHITE,
                "LBT Circle No": FILL_WHITE,
                "LBT Rate": FILL_WHITE,
                "Purchase Tax applicable": FILL_WHITE,
                "Total Bill Amount in Foreign Currency": FILL_WHITE,
                "Total Bill Amount in INR": FILL_WHITE,
                "Schedule Value": FILL_WHITE,
                "VAT": FILL_WHITE,
                "CST": FILL_WHITE,
                "C Form Applicable": FILL_WHITE,
                "Waybill Inward Number": FILL_WHITE,
                "Waybill Outward Number": FILL_WHITE,
                "Waybill Inward Counterfoil Received ": FILL_WHITE,
                "Waybill Outward Counterfoil Received ": FILL_WHITE,
                "Transporter Name": FILL_WHITE,
                "Transporter Address": FILL_WHITE,
                "Vehicle Number": FILL_WHITE,
                "SEZ": FILL_WHITE,
                "SEZ Amount": FILL_WHITE,
                "Import": FILL_WHITE,
                "Service Code": FILL_WHITE,
                "Service Tax Amount(If Applicable)": FILL_WHITE,
                "Others (Freight/ Octroi,etc)": FILL_WHITE,
                "Base Amt VAT": FILL_WHITE,
                "VAT %": FILL_WHITE,
                "Other Bill Component(Service/Freight)": FILL_WHITE,
                "VAT Rebate": FILL_WHITE,
                "WCT Base Value": FILL_WHITE,
                "CENVAT Amount passed on to customer": FILL_WHITE,
                "Retention": FILL_WHITE,
                "HSN Code": FILL_WHITE,
                "SAC Code": FILL_WHITE,
                "HSN SGST": FILL_WHITE,
                "HSN CGST": FILL_WHITE,
                "HSN IGST": FILL_WHITE,
                "SAC SGST": FILL_WHITE,
                "SAC CGST": FILL_WHITE,
                "SAC IGST": FILL_WHITE
            }

            for col in ws.columns:
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = 50 
                for cell in col:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for cell in ws[1]:
                header_text = str(cell.value).strip()
                
                if header_text in HEADER_COLOR_MAP:
                    cell.fill = HEADER_COLOR_MAP[header_text]
                else:
                    cell.fill = FILL_WHITE

            wb.save(file_path)
    except Exception as e:
        print(f"   > ⚠️ Formatting warning: {e}")

def safe_divide(value, divisor):
    if divisor <= 1:
        return value
    try:
        clean_val = str(value).replace(',', '').replace('₹', '').replace('INR', '').strip()
        float_val = float(clean_val)
        return round(float_val / divisor, 2)
    except (ValueError, TypeError):
        return value

def get_clean_float(value):
    if not value:
        return 0.0
    try:
        clean_val = str(value).replace(',', '').replace('₹', '').replace('INR', '').replace('USD', '').strip()
        return float(clean_val)
    except (ValueError, TypeError):
        return 0.0

def process_tags(raw_tags_input):
    """
    Helper to clean and validate tags.
    STRICT MODE: Only splits if input is a List or Stringified List.
    """
    if not raw_tags_input:
        return []

    initial_tags = []

    if isinstance(raw_tags_input, list):
        for t in raw_tags_input:
            initial_tags.append(str(t))
            
    elif isinstance(raw_tags_input, str):
        s = raw_tags_input.strip()
        
        if s.startswith('[') and s.endswith(']'):
            try:
                # Remove brackets
                content = s[1:-1]
                parts = content.split(',')
                for p in parts:
                    clean_p = p.strip().strip("'").strip('"')
                    if clean_p:
                        initial_tags.append(clean_p)
            except:
                initial_tags.append(s)
        else:
            if s:
                initial_tags.append(s)

    final_tags = []
    
    # FILTER LISTS
    BAD_KEYWORDS = {
        'VIDEO', 'AUDIO', 'CONF', 'KIT', 'UNIT', 'SYS', 'LIC', 'SUB', 'PRO', 'PLUS', 
        'WARRANTY', 'PRESENTATION', 'REPEATER', 'CABLE', 'HDMI', 
        'WINDOWS', 'SERVER', 'MICROSOFT', 'INTEL', 'PROCESSOR', 
        '7D76', '4X40', '8471', 'STANDARD', 'WIRELESS'
    }
    BRAND_PREFIXES = ('CISCO', 'LENOVO', 'HP', 'DELL', 'POLY', 'LOGI', 'APPLE')

    for t in initial_tags:
        clean_t = re.sub(r'[^a-zA-Z0-9-]', '', t.upper())
        length = len(clean_t)
        
        # VALIDATION GATE
        if length < 5: continue
        
        is_bad = False
        for bad in BAD_KEYWORDS:
            if bad in clean_t:
                is_bad = True
                break
        if is_bad: continue

        if clean_t.startswith(BRAND_PREFIXES): continue
        
        final_tags.append(clean_t)
            
    return list(dict.fromkeys(final_tags))

def flatten_invoice_data(data):
    """
    Revised logic: "SMART READER / DUMB BUILDER"
    """
    item_keys = ['Line Items', 'items', 'line_items']
    target_key = next((k for k in item_keys if k in data and isinstance(data[k], list)), None)
    items_list = data.get(target_key, [])
    
    global_tags = []
    possible_tag_keys = ['Asset Serial Number', 'Tag Nos', 'Serial No', 'Sr No']
    
    for pk in possible_tag_keys:
        if pk in data and data[pk]:
            global_tags = process_tags(data[pk])
            break 

    flattened_rows = []

    if not items_list:
        items_list = [data] 

    for i, item in enumerate(items_list):
        row_base = data.copy()
        if target_key in row_base: del row_base[target_key]
        
        if isinstance(item, dict):
            row_base.update(item)
            
        item_tags = []
        
        for pk in possible_tag_keys:
            if pk in item and item[pk]:
                item_tags.extend(process_tags(item[pk]))
        
        desc = item.get('Asset Description') or item.get('Description') or ''
        
        if str(desc).strip().startswith('['):
             item_tags.extend(process_tags(str(desc)))

        item_tags = list(dict.fromkeys(item_tags))

        if not item_tags and global_tags and i == 0:
            item_tags = global_tags
        if item_tags:
            divisor = len(item_tags)
            q_val = item.get('Quantity') or item.get('Qty')
            extracted_qty = get_clean_float(q_val)
            if extracted_qty > 0:
                divisor = extracted_qty
            
            for tag in item_tags:
                new_row = row_base.copy()
                new_row['Asset Serial Number'] = tag
                
                proration_keys = [
                    "Total Value", "Total Purchase Price", "Total Amount", "Grand Total", 
                    "Sub Total", "Subtotal", "IGST Amount", "CGST Amount", "SGST Amount", 
                    "Total SGST", "Total CGST", "Total IGST", "Tax Total", "Total Tax Amount", 
                    "Net Amount", "Tax Amount", "Rate", "Amount", "Value (INR)", "Total Base Price",
                    "IGST", "CGST", "SGST", "Total Tax", "HSN IGST", "Base Inv Amt (Excl Tax) - Labour", "Base Inv Amt (Excl Tax) - Material"
                ]
                for key, val in new_row.items():
                    if any(k.lower() == key.lower() for k in proration_keys):
                        new_row[key] = safe_divide(val, divisor)
                
                if 'Net Amount' in new_row: new_row['Unit Price'] = new_row['Net Amount']
                elif 'Total Base Price' in new_row: new_row['Unit Price'] = new_row['Total Base Price']
                
                new_row['Quantity'] = 1
                new_row['Qty'] = 1
                flattened_rows.append(new_row)
        else:
            new_row = row_base.copy()
            if 'Qty' not in new_row:
                new_row['Qty'] = new_row.get('Quantity', 1)
            flattened_rows.append(new_row)

    # 4. Final Metadata
    current_date_str = datetime.now().strftime("%Y-%m-%d")

    for row in flattened_rows:
        row['Date of Entry'] = current_date_str
        row['Data Entry Done By'] = "Auto"
        
        uom_val = row.get('UOM')
        if not uom_val or str(uom_val).strip() == "":
            row['UOM'] = "NUMBER"

        currency_val = row.get('Currency Code')
        currency_str = str(currency_val).strip().upper() if currency_val is not None else ""
        
        if not currency_str or 'INR' in currency_str or currency_str == 'NONE':
            row['Total Bill Amount in INR'] = 1
            row['Total Bill Amount in Foreign Currency'] = 0
        else:
            row['Total Bill Amount in INR'] = 0
            row['Total Bill Amount in Foreign Currency'] = 1
            
        hsn_sac_val = str(row.get('HSN/SAC') or row.get('HSN Code') or row.get('SAC Code') or row.get('Code') or '').strip()
        base_amt = row.get('Total Base Price') or row.get('Net Amount') or row.get('Unit Price') or ""
        
        if hsn_sac_val:
            if hsn_sac_val.startswith("99"):
                row['HSN/SAC'] = "SAC"
                row['Code'] = hsn_sac_val
                row['Base Inv Amt (Excl Tax) - Labour'] = base_amt
                row['Base Inv Amt (Excl Tax) - Material'] = ""
            else:
                row['HSN/SAC'] = "HSN"
                row['Code'] = hsn_sac_val
                row['Base Inv Amt (Excl Tax) - Material'] = base_amt
                row['Base Inv Amt (Excl Tax) - Labour'] = ""
        else:
            row['HSN/SAC'] = ""
            row['Code'] = ""
        
        po_num = str(row.get('PO Number', '')).strip()
        if po_num and not row.get('Client Code'):
            parts = po_num.split('/')
            if len(parts) > 1:
                row['Client Code'] = parts[0].strip()

        address = str(row.get('Deliver Address', '')).strip()
        current_loc = str(row.get('Delivery Location', '')).strip()
        if address and not current_loc:
            match = re.search(r'([\w\s]+?)[,\s-]*\d{6}', address)
            if match:
                loc_text = match.group(1).strip()
                loc_parts = loc_text.split(',')
                row['Delivery Location'] = loc_parts[-1].strip()

        base_price = get_clean_float(row.get('Total Base Price') or row.get('Net Amount'))
        total_tax = get_clean_float(row.get('Total Tax') or row.get('Tax Amount'))
        
        igst = get_clean_float(row.get('IGST') or row.get('IGST Amount'))
        cgst = get_clean_float(row.get('CGST') or row.get('CGST Amount'))
        sgst = get_clean_float(row.get('SGST') or row.get('SGST Amount'))

        if total_tax == 0 and (igst > 0 or cgst > 0 or sgst > 0):
            total_tax = igst + cgst + sgst
            row['Total Tax'] = round(total_tax, 2)

        if base_price > 0:
            grand_total = base_price + total_tax
            row['Total Purchase Price'] = round(grand_total, 2)
            if 'Grand Total' in row: row['Grand Total'] = round(grand_total, 2)
            if 'Total Amount' in row: row['Total Amount'] = round(grand_total, 2)

    return flattened_rows

def update_excel_sheet(vendor_name, data, file_path=EXCEL_FILE):
    if isinstance(data, list) and len(data) > 0:
        data_to_process = data
        if vendor_name == "Unknown_Vendor":
            vendor_name = data[0].get("Vendor Name", "Unknown_Vendor")
    elif isinstance(data, dict):
        data_to_process = [data]
    else:
        return [] 

    sheet_name = sanitize_sheet_name(vendor_name)
    all_rows = []
    
    for item in data_to_process:
        item['Vendor Name'] = vendor_name 
        extracted_rows = flatten_invoice_data(item)
        all_rows.extend(extracted_rows)

    if not all_rows:
        return []

    new_data_df = pd.DataFrame(all_rows)
    target_schema = load_schemas()
    
    if target_schema:
        final_df = pd.DataFrame(columns=target_schema)
        for col in target_schema:
            if col in new_data_df.columns:
                final_df[col] = new_data_df[col]
            else:
                final_df[col] = ""
        final_df = final_df.fillna("")
    else:
        final_df = new_data_df.fillna("")
        cols = list(final_df.columns)
        if 'Vendor Name' in cols:
            cols.insert(0, cols.pop(cols.index('Vendor Name')))
        final_df = final_df[cols]

    try:
        if not os.path.exists(file_path):
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                final_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"   > 🆕 Created new Excel file: {sheet_name}")
            columns = list(final_df.columns)
        else:
            xls = pd.ExcelFile(file_path)
            sheet_exists = sheet_name in xls.sheet_names
            
            if sheet_exists:
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                if target_schema:
                    existing_df = existing_df.reindex(columns=target_schema, fill_value="")
                
                blank_row = pd.DataFrame([[""] * len(final_df.columns)], columns=final_df.columns)
                updated_df = pd.concat([existing_df, blank_row, final_df], ignore_index=True)
                updated_df = updated_df.fillna("")
                
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                columns = list(updated_df.columns)
            else:
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                    final_df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"   > 🆕 Added new sheet: {sheet_name}")
                columns = list(final_df.columns)

        format_excel_sheet(file_path, sheet_name)
        return columns

    except Exception as e:
        print(f"   > ❌ Error updating Excel sheet '{sheet_name}': {e}")
        return []